import os
import tempfile
import shutil
from custom_mcp.excel_code_runner_mcp import excel_to_str_repr, run_excel_code_impl

TEST_XLSX = "tests/test_inputs/test1.xlsx"

def test_excel_to_str_repr_basic():
    output = excel_to_str_repr(TEST_XLSX)
    assert "Sheet Name: Sheet1" in output
    assert "Sheet Name: Sheet2" in output
    # Check for some known cell values and styles
    assert "D3: 1" in output
    assert "H3: D [bg:FFFFFF00]" in output
    assert "D6: =SUM(D3:D5) -> 2" in output

def test_run_excel_code_impl_success():
    # Copy input to a temp file for isolation
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = os.path.join(tmpdir, "input.xlsx")
        output_path = os.path.join(tmpdir, "output.xlsx")
        shutil.copyfile(TEST_XLSX, input_path)
        # User code: set D3 to 10 and save to output_path
        user_code = f'''
from openpyxl import load_workbook
wb = load_workbook("{input_path}")
ws = wb["Sheet1"]
ws["D3"].value = 10
wb.save("{output_path}")
wb.close()
'''
        result = run_excel_code_impl(user_code, output_path)
        assert "D3: 10" in result

def test_run_excel_code_impl_error():
    with tempfile.TemporaryDirectory() as tmpdir:
        output_path = os.path.join(tmpdir, "output.xlsx")
        user_code = "raise ValueError('test error')"
        result = run_excel_code_impl(user_code, output_path)
        assert result.startswith("ERROR:")
        assert "test error" in result
