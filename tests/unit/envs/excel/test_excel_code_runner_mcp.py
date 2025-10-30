from contextlib import contextmanager
import os
from pathlib import Path
import pytest
import shutil
from benchmax.envs.excel.workdir.excel_code_runner_mcp import run_excel_code_impl
from benchmax.envs.excel.workdir.excel_utils import excel_to_str_repr


# Fixtures
@pytest.fixture(scope="session")
def test_xlsx_path() -> str:
    return str(Path(__file__).parent / "test_inputs" / "test.xlsx")


@contextmanager
def temporary_cwd(path: Path):
    old_cwd = os.getcwd()
    os.chdir(path)
    try:
        yield
    finally:
        os.chdir(old_cwd)


def test_excel_to_str_repr_basic(test_xlsx_path: str):
    output = excel_to_str_repr(test_xlsx_path)
    assert "Sheet Name: Sheet1" in output
    assert "Sheet Name: Sheet2" in output
    # Check for some known cell values and styles
    assert "D3: 1" in output
    assert "H3: D" in output
    assert "D6: =SUM(D3:D5) -> 2" in output


def test_run_excel_code_impl_success(tmp_path: Path, test_xlsx_path: str):
    # Copy input to a temp file for isolation
    input_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    shutil.copyfile(test_xlsx_path, input_path)

    # User code: set D3 to 10 and save to output_path
    user_code = f'''
from openpyxl import load_workbook
wb = load_workbook("{input_path}")
ws = wb["Sheet1"]
ws["D3"].value = 10
wb.save("{output_path}")
wb.close()
'''

    # Run code with temporary cwd
    with temporary_cwd(tmp_path):
        result = run_excel_code_impl(user_code, "output.xlsx")

    assert "D3: 10" in result


def test_run_excel_code_impl_error(tmp_path: Path):
    output_path = tmp_path / "output.xlsx"
    user_code = "raise ValueError('test error')"

    with temporary_cwd(tmp_path):
        result = run_excel_code_impl(user_code, str(output_path))

    assert result.startswith("ERROR:")
    assert "test error" in result
