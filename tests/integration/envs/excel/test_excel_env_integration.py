"""
Integration and end-to-end tests for ExcelEnv.

"""

import asyncio
import importlib.util
import random
import pytest
import uuid
import json
from pathlib import Path
from datasets import Dataset
from typing import AsyncGenerator, Dict, List, Tuple, Any

from benchmax.envs.excel.excel_env import ExcelEnv
from benchmax.envs.excel.data_utils import download_and_extract
from benchmax.envs.mcp.provisioners.local_provisioner import LocalProvisioner
from benchmax.envs.mcp.provisioners.manual_provisioner import ManualProvisioner


# Fixtures
@pytest.fixture(scope="session")
def excel_workdir() -> Path:
    """Path to excel MCP workdir."""
    spec = importlib.util.find_spec("benchmax.envs.excel")
    if not spec or not spec.submodule_search_locations:
        raise RuntimeError("Could not locate benchmax.envs.excel package")

    excel_pkg_dir = Path(spec.submodule_search_locations[0])
    workdir = excel_pkg_dir / "workdir"

    if not workdir.exists():
        raise FileNotFoundError(f"Expected workdir not found at: {workdir}")

    return workdir


@pytest.fixture(scope="session")
async def local_excel_servers(
    excel_workdir: Path,
) -> AsyncGenerator[Tuple[List[str], str], None]:
    """
    Provision local Excel servers once for entire test session.
    Returns (addresses, api_secret) tuple.
    """
    api_secret = uuid.uuid4().hex
    provisioner = LocalProvisioner(
        workdir_path=excel_workdir,
        num_servers=4,
        base_port=8090,
    )

    addresses = await provisioner.provision_servers(api_secret)

    yield addresses, api_secret

    await provisioner.teardown()


@pytest.fixture
async def excel_env(
    local_excel_servers: Tuple[List[str], str], 
    excel_workdir: Path,
    excel_dataset: Tuple[Dataset, str]
) -> AsyncGenerator[ExcelEnv, None]:
    """
    Create fresh ExcelEnv using reused local servers and test dataset path.
    Each test gets clean env instance.
    """
    addresses, api_secret = local_excel_servers
    _, dataset_path = excel_dataset

    manual_provisioner = ManualProvisioner(addresses)
    env = ExcelEnv(
        workdir_path=excel_workdir,
        provisioner=manual_provisioner,
        api_secret=api_secret,
        provision_at_init=True,
        dataset_path=dataset_path,
    )

    yield env

    await env.shutdown()


@pytest.fixture(scope="session")
def excel_dataset(session_tmp_path: Path) -> Tuple[Dataset, str]:
    """Load SpreadsheetBench sample dataset once for entire test session."""
    # Use temporary directory for test data
    test_data_path = session_tmp_path / "excel_test_data"
    
    # Download sample dataset
    folder_path = Path(test_data_path) / "sample_data_200"
    json_path = folder_path / "dataset.json"
    
    if not json_path.exists():
        download_and_extract(
            "https://github.com/RUCKBReasoning/SpreadsheetBench/raw/refs/heads/main/data/sample_data_200.tar.gz",
            test_data_path,
        )
    
    # Load dataset
    with open(json_path, "r") as f:
        data = json.load(f)
        for example in data:
            example["id"] = str(example["id"])  # Ensure IDs are strings
    
    dataset = Dataset.from_list(data)
    
    return dataset, str(folder_path)


@pytest.fixture
def sample_dataset_example(excel_dataset: Tuple[Dataset, str]) -> Dict[str, Any]:
    """Return a random example from the dataset."""
    dataset, _ = excel_dataset
    idx = random.randint(0, len(dataset) - 1)
    return dataset[idx]


class TestExcelDataset:
    @pytest.mark.slow
    @pytest.mark.excel
    def test_load_dataset(self, excel_dataset: Tuple[Dataset, str]) -> None:
        """Verify dataset loads correctly with all required fields."""
        dataset, base_dir = excel_dataset
        
        # Verify minimum row count
        assert len(dataset) > 0, "Dataset should not be empty"
        
        # Check required fields in dataset
        first_example = dataset[0]
        required_fields = ["id", "spreadsheet_path", "instruction", "instruction_type", "answer_position"]
        for field in required_fields:
            assert field in first_example, f"Dataset should have '{field}' field"
        
        # Verify spreadsheet files exist
        spreadsheet_path = first_example["spreadsheet_path"]
        example_id = first_example["id"]
        input_file = Path(base_dir) / spreadsheet_path / f"1_{example_id}_input.xlsx"
        assert input_file.exists(), f"Input spreadsheet should exist at {input_file}"

    @pytest.mark.slow
    @pytest.mark.excel
    def test_dataset_preprocess(self, excel_dataset: Tuple[Dataset, str]) -> None:
        """Verify dataset preprocessing produces valid ExcelExample."""
        dataset, dataset_path = excel_dataset
        first_example = dataset[0]
        
        standardized = ExcelEnv.dataset_preprocess(first_example, dataset_path=dataset_path)
        
        # Check StandardizedExample fields
        assert standardized["prompt"] is not None
        assert "Instruction:" in standardized["prompt"]
        assert "Spreadsheet Path:" in standardized["prompt"]
        assert "Output Path:" in standardized["prompt"]
        
        # Check ExcelExample specific fields
        assert standardized["id"] == str(first_example["id"])
        assert standardized["answer_position"] == first_example["answer_position"]
        assert standardized["output_filename"].endswith(".xlsx")
        assert standardized["ground_truth_filename"].endswith(".xlsx")
        assert standardized["init_rollout_args"] is not None
        assert "input_src_path" in standardized["init_rollout_args"]


class TestExcelTools:
    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_list_tools(self, excel_env: ExcelEnv) -> None:
        """Verify Excel manipulation tools are in tool list."""
        tools = await excel_env.list_tools()
        tool_names = [tool.name for tool in tools]
        
        # Check for simple tool presence
        assert "run_excel_code" in tool_names, "Should have 'run_excel_code' tool available"

    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_run_excel_code_tool(
        self, excel_env: ExcelEnv, excel_dataset: Tuple[Dataset, str], unique_rollout_id: str
    ) -> None:
        """Test actual tool call to run_excel_code."""
        dataset, dataset_path = excel_dataset
        first_example = dataset[0]
        
        standardized = ExcelEnv.dataset_preprocess(first_example, dataset_path=dataset_path)
        
        # Initialize rollout
        rollout_args = standardized.get("init_rollout_args") or {}
        await excel_env.init_rollout(unique_rollout_id, **rollout_args)
        
        # Create simple Python code to read and modify the Excel file
        python_code = f"""
from openpyxl import load_workbook

# Load the input workbook
wb = load_workbook('{standardized["output_filename"].replace("output", "input")}')
ws = wb.active

# Make a simple modification (e.g., set a cell value)
ws['A1'] = 'Test Value'

# Save to output path
wb.save('{standardized["output_filename"]}')
wb.close()
"""
        
        # Run the tool
        result = await excel_env.run_tool(
            unique_rollout_id,
            "run_excel_code",
            python_code=python_code,
            output_excel_path=standardized["output_filename"]
        )
        
        # Verify result is returned (should be string representation of Excel or error)
        assert isinstance(result, str), "Tool should return a string"
        assert len(result) > 0, "Tool result should not be empty"
        
        # Clean up
        await excel_env.release_rollout(unique_rollout_id)


class TestExcelWorkspace:
    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_init_rollout_copies_file(
        self, excel_env: ExcelEnv, excel_dataset: Tuple[Dataset, str], unique_rollout_id: str, tmp_path: Path
    ) -> None:
        """Verify input spreadsheet is copied to workspace during init_rollout."""
        dataset, dataset_path = excel_dataset
        first_example = dataset[0]
        
        standardized = ExcelEnv.dataset_preprocess(first_example, dataset_path=dataset_path)
        
        # Initialize rollout
        rollout_args = standardized["init_rollout_args"] or {}
        await excel_env.init_rollout(unique_rollout_id, **rollout_args)
        
        # Copy file from workspace to verify it exists and matches original
        dest_path = tmp_path / "copied_input.xlsx"
        input_src_path = Path(rollout_args["input_src_path"])
        
        # Copy from workspace
        await excel_env.copy_from_workspace(unique_rollout_id, input_src_path.name, dest_path)
        
        # Verify file was copied
        assert dest_path.exists(), f"File should be copied from workspace to {dest_path}"
        
        # Verify file size matches (basic check)
        original_size = input_src_path.stat().st_size
        copied_size = dest_path.stat().st_size
        assert copied_size == original_size, "Copied file should have same size as original"
        
        # Clean up
        await excel_env.release_rollout(unique_rollout_id)

    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_workspace_isolation(
        self, excel_env: ExcelEnv, excel_dataset: Tuple[Dataset, str], unique_rollout_id: str, tmp_path: Path
    ) -> None:
        """Verify multiple rollouts have isolated workspaces."""
        dataset, dataset_path = excel_dataset
        
        # Get two different examples
        example1 = dataset[0]
        example2 = dataset[min(1, len(dataset) - 1)]
        
        standardized1 = ExcelEnv.dataset_preprocess(example1, dataset_path=dataset_path)
        standardized2 = ExcelEnv.dataset_preprocess(example2, dataset_path=dataset_path)

        rollout_id1 = f"{unique_rollout_id}-1"
        rollout_id2 = f"{unique_rollout_id}-2"

        # Initialize both rollouts
        rollout_args1 = standardized1.get("init_rollout_args") or {}
        rollout_args2 = standardized2.get("init_rollout_args") or {}

        await excel_env.init_rollout(rollout_id1, **rollout_args1)
        await excel_env.init_rollout(rollout_id2, **rollout_args2)

        # Create subdirectories in tmp_path for each workspace
        temp_dir1 = tmp_path / "workspace1"
        temp_dir2 = tmp_path / "workspace2"
        temp_dir1.mkdir()
        temp_dir2.mkdir()
        
        # Get filenames
        src_path1 = Path(rollout_args1["input_src_path"])
        src_path2 = Path(rollout_args2["input_src_path"])

        input_filename1 = src_path1.name
        input_filename2 = src_path2.name

        # Copy from each workspace
        dest1 = temp_dir1 / "workspace1_file.xlsx"
        dest2 = temp_dir2 / "workspace2_file.xlsx"
        
        await excel_env.copy_from_workspace(rollout_id1, input_filename1, dest1)
        await excel_env.copy_from_workspace(rollout_id2, input_filename2, dest2)
        
        # Verify both files exist
        assert dest1.exists(), "File from workspace 1 should exist"
        assert dest2.exists(), "File from workspace 2 should exist"
        
        # Verify files match their original sources
        original_size1 = src_path1.stat().st_size
        original_size2 = src_path2.stat().st_size

        assert dest1.stat().st_size == original_size1, "Workspace 1 file should match original 1"
        assert dest2.stat().st_size == original_size2, "Workspace 2 file should match original 2"
        
        # If the examples are different, verify the files are different
        if example1["id"] != example2["id"]:
            assert dest1.stat().st_size != dest2.stat().st_size or \
                   dest1.read_bytes() != dest2.read_bytes(), \
                   "Different examples should produce different workspace files"
        
        # Verify that file from workspace1 is NOT accessible from workspace2 by trying to copy
        # the other workspace's file (should fail or not exist)
        try:
            temp_cross = temp_dir1 / "should_not_exist.xlsx"
            await excel_env.copy_from_workspace(rollout_id1, input_filename2, temp_cross)
            # If we get here and files are different, this is a problem
            if example1["id"] != example2["id"]:
                pytest.fail("Should not be able to access file from different workspace")
        except Exception:
            # Expected - file doesn't exist in workspace1
            pass
        
        # Clean up
        await excel_env.release_rollout(rollout_id1)
        await excel_env.release_rollout(rollout_id2)


class TestExcelReward:
    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_reward_missing_output(
        self, excel_env: ExcelEnv, excel_dataset: Tuple[Dataset, str], unique_rollout_id: str
    ) -> None:
        """Test reward when output file doesn't exist."""
        dataset, dataset_path = excel_dataset
        first_example = dataset[0]
        
        standardized = ExcelEnv.dataset_preprocess(first_example, dataset_path=dataset_path)
        
        rollout_args = standardized.get("init_rollout_args") or {}
        await excel_env.init_rollout(unique_rollout_id, **rollout_args)
        
        # Compute reward without creating output file
        rewards = await excel_env.compute_reward(
            unique_rollout_id,
            completion="Task completed.",
            ground_truth="",
            answer_position=standardized["answer_position"],
            output_filename=standardized["output_filename"],
            ground_truth_filename=standardized["ground_truth_filename"],
            spreadsheet_base_dir=standardized["spreadsheet_base_dir"],
        )
        
        # When output file doesn't exist, reward should be zero
        assert "spreadsheet" in rewards
        assert rewards["spreadsheet"] == 0.0

    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_reward_with_matching_output(
        self, excel_env: ExcelEnv, excel_dataset: Tuple[Dataset, str], unique_rollout_id: str
    ) -> None:
        """Test reward computation with output that matches ground truth."""
        dataset, dataset_path = excel_dataset
        first_example = dataset[0]
        
        standardized = ExcelEnv.dataset_preprocess(first_example, dataset_path=dataset_path)
        
        rollout_args = standardized.get("init_rollout_args") or {}
        await excel_env.init_rollout(unique_rollout_id, **rollout_args)
        
        # Copy the ground truth file as the output (to simulate perfect completion)
        ground_truth_src = Path(standardized["spreadsheet_base_dir"]) / standardized["ground_truth_filename"]
        await excel_env.copy_to_workspace(unique_rollout_id, ground_truth_src, standardized["output_filename"])
        
        # Compute reward
        rewards = await excel_env.compute_reward(
            unique_rollout_id,
            completion="Task completed successfully.",
            ground_truth="",
            answer_position=standardized["answer_position"],
            output_filename=standardized["output_filename"],
            ground_truth_filename=standardized["ground_truth_filename"],
            spreadsheet_base_dir=standardized["spreadsheet_base_dir"],
        )
        
        # With matching output, reward should be 1.0
        assert "spreadsheet" in rewards
        assert rewards["spreadsheet"] == 1.0

    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_reward_with_mismatched_output(
        self, excel_env: ExcelEnv, excel_dataset: Tuple[Dataset, str], unique_rollout_id: str
    ) -> None:
        """Test reward computation with output that doesn't match ground truth."""
        dataset, dataset_path = excel_dataset
        
        # Get two different examples to ensure mismatch
        if len(dataset) < 2:
            pytest.skip("Dataset needs at least 2 examples for mismatch test")
        
        first_example = dataset[0]
        second_example = dataset[1]
        
        standardized = ExcelEnv.dataset_preprocess(first_example, dataset_path=dataset_path)
        standardized_second = ExcelEnv.dataset_preprocess(second_example, dataset_path=dataset_path)
        
        rollout_args = standardized.get("init_rollout_args") or {}
        await excel_env.init_rollout(unique_rollout_id, **rollout_args)
        
        # Copy output from second example (different from ground truth)
        output_src = Path(standardized_second["spreadsheet_base_dir"]) / standardized_second["ground_truth_filename"]
        await excel_env.copy_to_workspace(unique_rollout_id, output_src, standardized["output_filename"])
        
        # Compute reward
        rewards = await excel_env.compute_reward(
            unique_rollout_id,
            completion="Task completed.",
            ground_truth="",
            answer_position=standardized["answer_position"],
            output_filename=standardized["output_filename"],
            ground_truth_filename=standardized["ground_truth_filename"],
            spreadsheet_base_dir=standardized["spreadsheet_base_dir"],
        )
        
        # With mismatched output, reward should be between 0.0 and 1.0 (likely 0.0)
        assert "spreadsheet" in rewards
        assert 0.0 <= rewards["spreadsheet"] < 1.0


class TestExcelEndToEnd:
    """End-to-end tests for full ExcelEnv workflows."""

    async def _run_single_rollout(
        self,
        rollout_id: str,
        env: ExcelEnv,
        example: Dict[str, Any],
        dataset_path: str
    ) -> Dict[str, float]:
        """
        Execute full rollout workflow:
        1. Preprocess example
        2. Init rollout (copies input file to workspace)
        3. Optionally run tool
        4. Compute reward with dummy completion

        Returns reward dict.
        """
        # Preprocess
        standardized = env.dataset_preprocess(example, dataset_path=dataset_path)

        # Init rollout
        rollout_args = standardized.get("init_rollout_args") or {}
        await env.init_rollout(rollout_id, **rollout_args)

        # Run tool
        # Create simple Python code that copies input to output
        input_filename = Path(rollout_args["input_src_path"]).name
        python_code = f"""
from openpyxl import load_workbook

# Load the input workbook
wb = load_workbook('{input_filename}')

# Save to output path (no modifications for this test)
wb.save('{standardized["output_filename"]}')
wb.close()
"""
        await env.run_tool(
            rollout_id,
            "run_excel_code",
            python_code=python_code,
            output_excel_path=standardized["output_filename"]
        )

        # Compute reward
        rewards = await env.compute_reward(
            rollout_id,
            completion="Task completed.",
            ground_truth="",
            answer_position=standardized["answer_position"],
            output_filename=standardized["output_filename"],
            ground_truth_filename=standardized["ground_truth_filename"],
            spreadsheet_base_dir=standardized["spreadsheet_base_dir"],
        )

        return rewards

    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_rollout_basic_workflow(
        self, excel_env: ExcelEnv, excel_dataset: Tuple[Dataset, str]
    ) -> None:
        """Full rollout with basic workflow."""
        dataset, dataset_path = excel_dataset
        first_example = dataset[0]
        
        reward = await self._run_single_rollout(
            rollout_id="test-rollout-basic",
            env=excel_env,
            example=first_example,
            dataset_path=dataset_path
        )

        assert "spreadsheet" in reward
        assert reward["spreadsheet"] == 0

    @pytest.mark.slow
    @pytest.mark.excel
    @pytest.mark.asyncio
    async def test_parallel_rollouts(
        self, excel_env: ExcelEnv, excel_dataset: Tuple[Dataset, str]
    ) -> None:
        """Multiple rollouts should execute correctly in parallel."""
        dataset, dataset_path = excel_dataset
        
        # Get three examples
        num_examples = min(3, len(dataset))
        examples = [dataset[i] for i in range(num_examples)]
        
        results = await asyncio.gather(
            *[
                self._run_single_rollout(
                    rollout_id=f"test-rollout-parallel-{i}",
                    env=excel_env,
                    example=examples[i],
                    dataset_path=dataset_path,
                )
                for i in range(num_examples)
            ]
        )

        # All should return reward dicts
        assert len(results) == num_examples
        for reward in results:
            assert "spreadsheet" in reward
            assert reward["spreadsheet"] == 0