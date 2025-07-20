import pytest
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from benchmax.envs.excel.excel_utils import compare_excel_cells, evaluate_excel, excel_to_str_repr

@pytest.fixture(scope="session")
def setup_files():
    base_excel_path = Path("tests/test_inputs/test1.xlsx")
    gt_path = Path("tests/test_inputs/test1_gt.xlsx")
    output_path = Path("tests/test_inputs/test1_output.xlsx")

    # Create ground truth file
    wb_gt = load_workbook(base_excel_path)
    ws_gt = wb_gt.active
    assert ws_gt is not None, "Worksheet should not be None"
    ws_gt["H3"] = "Text Value"
    ws_gt["H4"] = 4
    ws_gt["H5"] = 2  # Numeric value
    ws_gt["H6"] = "Mismatch Value"
    ws_gt["I3"] = "Matching Value"
    ws_gt["I4"] = "Matching Value"
    ws_gt["I3"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    ws_gt["I4"].font = Font(color="FF0000")  # Red font
    wb_gt.save(gt_path)
    wb_gt.close()

    # Create output file
    wb_output = load_workbook(base_excel_path)
    ws_output = wb_output.active
    assert ws_output is not None, "Worksheet should not be None"
    ws_output["H3"] = "Text Value"
    ws_output["H4"] = 4
    ws_output["H5"] = "=G5"  # Formula
    ws_output["H6"] = "Different Mismatch Value"
    ws_output["I3"] = "Matching Value"
    ws_output["I4"] = "Matching Value"
    ws_output["I3"].fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
    ws_output["I4"].font = Font(color="0000FF")  # Blue font
    wb_output.save(output_path)
    wb_output.close()

    evaluate_excel(str(gt_path))
    evaluate_excel(str(output_path))

    yield gt_path, output_path

    # Cleanup - this will run once at the end of the session
    if gt_path.exists():
        gt_path.unlink()
    if output_path.exists():
        output_path.unlink()

# Test for mismatched values in a single cell
def test_single_cell_comparison_mismatch(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "H6"
    )
    assert not result
    assert "Value mismatch" in message

# Test for matching values in a single cell
def test_single_cell_comparison_match(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "H3"
    )
    assert result

# Test for mismatched values in a range of cells
def test_range_comparison_mismatch(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "D4:H6"
    )
    assert not result
    assert "Value mismatch" in message

# Test for matching values in a range of cells
def test_range_comparison_match(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "H5:H5"
    )
    assert result

# Test for handling sheet names with mismatched values
def test_sheet_name_handling_mismatch(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "'Sheet1'!H6"
    )
    assert not result
    assert "Value mismatch" in message

# Test for handling sheet names with matching values
def test_sheet_name_handling_match(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "'Sheet1'!H3"
    )
    assert result

# Test for comparing cell formatting with mismatched styles
def test_fill_color_comparison_mismatch(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "I3",
        is_CF=True
    )
    assert not result
    assert "Fill color mismatch" in message


# Test for comparing cell formatting with mismatched styles
def test_font_color_comparison_mismatch(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "I4",
        is_CF=True
    )
    assert not result
    assert "Font color mismatch" in message

# Test for comparing cell formatting with matching styles
def test_formatting_comparison_match(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "H3",
        is_CF=True
    )
    assert result

# Test for handling missing sheets with mismatched values
def test_missing_sheet_mismatch(setup_files):
    ground_truth_path, test_output_path = setup_files
    result, message = compare_excel_cells(
        str(ground_truth_path),
        str(test_output_path),
        "'NonExistentSheet'!A1"
    )
    assert not result
    assert "Worksheet 'NonExistentSheet' not found" in message